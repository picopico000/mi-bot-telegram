#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import datetime
import logging

# Para leer variables desde un archivo .env (opcional)
from dotenv import load_dotenv

# Librerías de python-telegram-bot v13 (API antigua con Updater)
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    Filters,
    CallbackContext
)
from telegram import Update

# Para manejar Excel
from openpyxl import Workbook, load_workbook

# =========== 1. CONFIGURACIÓN DE LOGGING ===========
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# =========== 2. CARGA DEL TOKEN ===========
load_dotenv()  # Cargamos variables del archivo .env (si existe)
TOKEN = "7843970894:AAFBEaBejnRvPWHGKI5gdTMrcSuhHOZoeBo"

# Si prefieres poner el token directo, comenta la línea anterior y haz:
# TOKEN = "123456:ABC-DEF..."

if not TOKEN:
    raise ValueError("No se encontró el TOKEN. Configura .env o ponlo directo en el código.")

# =========== 3. ARCHIVO EXCEL ===========
EXCEL_FILE = "DATA.xlsx"

def get_workbook():
    """
    Carga (o crea) el archivo Excel con las columnas que necesitamos.
    """
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "DatosCripto"
        headers = [
            "Fecha/Hora",
            "Token",        # ← capturaremos aquí (ej: POSTIZ)
            "Address",
            "Volume",
            "RATING",
            "Market_Cap",
            "Distribution",
            "DevHold",
            "DevHoldValue",
            "test",
            "Holders",
            "Insider",
            "Normal",
            "Skizo",
            "Risk"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)
        logger.info("Creado el archivo Excel con encabezados: %s", headers)
    else:
        wb = load_workbook(EXCEL_FILE)
    return wb

def agregar_a_excel(datos):
    """
    Agrega la lista 'datos' como una nueva fila a la hoja 'DatosCripto'.
    """
    wb = get_workbook()
    ws = wb["DatosCripto"]
    ws.append(datos)
    wb.save(EXCEL_FILE)
    logger.info("Datos agregados al Excel: %s", datos)

# =========== 4. PARSEO DE MENSAJES ===========
def parse_message(text: str):
    """
    Busca en el texto:
      - Token al inicio de línea con '$' (ej: "$POSTIZ"), permitiendo espacios: ^[ \t]*\$
      - Address: cualquier texto después de "Address:"
      - Market_Cap: valor después de "Market Cap:"
      - #Volume ...
      - #RATING: ...
      - #Distribution: ...
      - #DevHold: ...
      - #DevHoldValue: ...
      - #test: ...
      - #Holders: ...
      - #Insider: ...
      - #Normal: ...
      - #Skizo: ...
    """
    import re
    import datetime

     # Capturar la fecha y hora actual
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Token al inicio de la línea con '$'
    token_match = re.search(r'^\s*\$\s*([^\s]+)', text, re.MULTILINE)
    token = token_match.group(1) if token_match else ""

    risk_match = re.search(r'#Risk\b', text)
    risk = "Risk" if risk_match else "Normal"

    # Address en formato general (después de "Address:")
    address_match = re.search(r'Address:\s*([^\s]+)', text)
    address = address_match.group(1) if address_match else ""

    # Market Cap (después de "Market Cap:")
    market_cap_match = re.search(r'Market Cap:\s*\$([\d.,]+)', text)
    market_cap = market_cap_match.group(1) if market_cap_match else ""

    # RATING (después de "#RATING:")
    rating_match = re.search(r'#RATING:\s*([\d.]+)', text)
    rating = rating_match.group(1) if rating_match else ""

    # Volume
    volume_match = re.search(r'#Volume\s+([\d.]+)', text)
    volume = volume_match.group(1) if volume_match else ""

    # Distribution
    distribution_match = re.search(r'#Distribution:\s*([\d.]+)', text)
    distribution = distribution_match.group(1) if distribution_match else ""

    # DevHold
    devhold_match = re.search(r'#DevHold:\s*([\d.]+)', text)
    devhold = devhold_match.group(1) if devhold_match else ""

    # DevHoldValue
    devhold_value_match = re.search(r'#DevHoldValue:\s*\$?([\d.]+)', text)
    devhold_value = devhold_value_match.group(1) if devhold_value_match else ""

    # test
    test_match = re.search(r'#test:\s*(\S+)', text)
    test = test_match.group(1) if test_match else ""

    # Holders
    holders_match = re.search(r'#Holders:\s*(\S+)', text)
    holders = holders_match.group(1) if holders_match else ""

    # Insider
    insider_match = re.search(r'#Insider:\s*(\S+)', text)
    insider = insider_match.group(1) if insider_match else ""

    # Normal
    normal_match = re.search(r'#Normal:\s*(\S+)', text)
    normal = normal_match.group(1) if normal_match else ""

    # Skizo
    skizo_match = re.search(r'#Skizo:\s*(\S+)', text)
    skizo = skizo_match.group(1) if skizo_match else ""

    #Status
    

    """ #Status
    status_normal_match = re.search(r'#Normal:\s*(\S+)', text)
    status_normal = skizo_match.group(1) if skizo_match else ""
    """

    # Retornar todos los datos
    return [
        now,            # Fecha/Hora
        token,          # Token (ej: POSTIZ, PEPE)
        address,        # Address
        volume,         # Volume
        rating,         # RATING
        market_cap,     # Market Cap
        distribution,   # Distribution
        devhold,        # DevHold
        devhold_value,  # DevHoldValue
        test,           # test
        holders,        # Holders
        insider,        # Insider
        normal,         # Normal
        skizo,          # Skizo
        risk
        
    ]
    
# =========== 5. HANDLERS DE COMANDOS ===========
def start(update: Update, context: CallbackContext):
    user = update.effective_user
    update.message.reply_text(
        f"¡Hola, {user.first_name}! Reenvía un mensaje con #Volume, #RATING, etc.,\n"
        "y si tu token está en una línea con '$', por ejemplo:\n\n"
        "$POSTIZ\nAddress: 0x...\n#Volume 1234\n...\n\n"
        "Lo registraré todo en Excel."
    )

def help_command(update: Update, context: CallbackContext):
    help_text = (
        "Comandos:\n"
        "/start - Inicia la interacción\n"
        "/help - Muestra esta ayuda\n"
        "/done - Finaliza el proceso\n"
        "/cancel - Cancela cualquier flujo\n\n"
        "Reenvía un mensaje con:\n"
        "- $TOKEN (en su propia línea)\n"
        "- Address: 0x...\n"
        "- #Volume, #RATING, #Distribution, #DevHold, #DevHoldValue...\n"
        "y lo guardaré en el Excel."
    )
    update.message.reply_text(help_text)

def done(update: Update, context: CallbackContext):
    update.message.reply_text("¡Proceso finalizado! Si necesitas más datos, vuelve a reenviar o usa /start. MAN DAO POR CULOOO AYUDA SOY UN INDIO SECUESTRADO ESTOY EN EL SOTANO DE FRAN Y NO ME PARA DE DAR POR CULO")
    logger.info("Usuario %s ha usado /done", update.effective_user.username)

def cancel(update: Update, context: CallbackContext):
    update.message.reply_text("Operación cancelada. Si necesitas ayuda, usa /help.")
    logger.info("Usuario %s ha usado /cancel", update.effective_user.username)

# =========== 6. HANDLER PARA MENSAJES REENVIADOS ===========
def received_message(update: Update, context: CallbackContext):
    message = update.message
    if message.forward_date:
        text = message.text or message.caption or ""
        data_row = parse_message(text)
        agregar_a_excel(data_row)
        update.message.reply_text(
            "Datos guardados correctamente.\n"
            "¿Quieres seguir enviando más? Si has terminado, usa /done."
        )
    else:
        update.message.reply_text(
            "Este mensaje no parece estar reenviado.\n"
            "Por favor reenvía el mensaje que contenga #Volume, #RATING, etc., "
            "y en otra línea $TOKEN y 'Address: ...' si procede."
        )

# =========== 7. MANEJO GLOBAL DE ERRORES ===========
def error_handler(update: object, context: CallbackContext):
    logger.error(msg="Ocurrió un error:", exc_info=context.error)
    if update and hasattr(update, "message") and update.message:
        update.message.reply_text("Lo siento, ocurrió un error inesperado. Intenta de nuevo más tarde.")

# =========== 8. FUNCIÓN PRINCIPAL (MAIN) ===========
def main():
    if not TOKEN:
        raise ValueError("No se ha proporcionado un TOKEN válido.")

    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher

    # Comandos
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("help", help_command))
    dp.add_handler(CommandHandler("done", done))
    dp.add_handler(CommandHandler("cancel", cancel))

    # Mensajes reenviados
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command, received_message))

    # Errores globales
    dp.add_error_handler(error_handler)

    logger.info("Bot iniciado. Esperando mensajes...")
    updater.start_polling()
    updater.idle()

if __name__ == "__main__":
    main()
